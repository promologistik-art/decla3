from openpyxl import load_workbook
from typing import Dict
from config import TEMPLATES_DIR
from loguru import logger

class KUDIRGenerator:
    """Генератор КУДИР"""
    
    def __init__(self):
        self.template_path = TEMPLATES_DIR / "KUDIR_template.xlsx"
    
    def generate(self, data: Dict, output_path: str, inn: str, name: str):
        """Заполняет КУДИР"""
        try:
            wb = load_workbook(self.template_path)
            
            # Титульный лист
            ws_title = wb['Лист1']
            self._fill_title(ws_title, inn, name)
            
            # Раздел I (Доходы и расходы)
            ws_section1 = wb['Лист2']
            self._fill_section1(ws_section1, data)
            
            ws_section1_q3 = wb['Лист3']
            self._fill_section1_q3_q4(ws_section1_q3, data)
            
            wb.save(output_path)
            logger.info(f"КУДИР сохранён: {output_path}")
            
        except Exception as e:
            logger.error(f"Ошибка генерации КУДИР: {e}")
            raise
    
    def _fill_title(self, ws, inn: str, name: str):
        """Заполняет титульный лист"""
        # ИНН (поиск и заполнение)
        for row in ws.iter_rows(min_row=1, max_row=50):
            for cell in row:
                if cell.value and 'ИНН' in str(cell.value):
                    for i, digit in enumerate(inn[:12]):
                        ws.cell(row=cell.row, column=cell.column + i + 1).value = digit
                    break
        
        # Наименование
        ws['B10'] = name
    
    def _fill_section1(self, ws, data: Dict):
        """Заполняет раздел I (кварталы 1-2)"""
        # Итого за I квартал
        ws['D50'] = round(data.get('income_q1', 0), 2)
        ws['E50'] = round(data.get('expense_q1', 0), 2)
        
        # Итого за II квартал
        ws['D100'] = round(data.get('income_q2', 0), 2)
        ws['E100'] = round(data.get('expense_q2', 0), 2)
        
        # Итого за полугодие
        ws['D150'] = round(data.get('income_q1', 0) + data.get('income_q2', 0), 2)
        ws['E150'] = round(data.get('expense_q1', 0) + data.get('expense_q2', 0), 2)
    
    def _fill_section1_q3_q4(self, ws, data: Dict):
        """Заполняет раздел I (кварталы 3-4)"""
        # Итого за III квартал
        ws['D50'] = round(data.get('income_q3', 0), 2)
        ws['E50'] = round(data.get('expense_q3', 0), 2)
        
        # Итого за IV квартал
        ws['D100'] = round(data.get('income_q4', 0), 2)
        ws['E100'] = round(data.get('expense_q4', 0), 2)
        
        # Итого за год
        total_income = sum([data.get(f'income_{q}', 0) for q in ['q1', 'q2', 'q3', 'q4']])
        total_expense = sum([data.get(f'expense_{q}', 0) for q in ['q1', 'q2', 'q3', 'q4']])
        ws['D150'] = round(total_income, 2)
        ws['E150'] = round(total_expense, 2)