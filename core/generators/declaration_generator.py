from openpyxl import load_workbook
from typing import Dict
from config import TEMPLATES_DIR
from loguru import logger
import os

class DeclarationGenerator:
    """Генератор налоговой декларации УСН"""
    
    def __init__(self):
        self.template_path = TEMPLATES_DIR / "Declaration_template.xlsx"
    
    def generate(self, data: Dict, output_path: str, inn: str, name: str, 
                 oktmo: str = '36701000', tax_code: str = '6317'):
        """Заполняет налоговую декларацию"""
        try:
            wb = load_workbook(self.template_path)
            
            # Титульный лист
            ws_title = wb['Титул']
            self._fill_title(ws_title, inn, name, oktmo, tax_code)
            
            # Раздел 1.1 (для УСН 6%)
            ws_section1_1 = wb['Раздел 1.1']
            self._fill_section1_1(ws_section1_1, data, oktmo)
            
            # Раздел 2.1.1 (расчёт налога)
            ws_section2_1_1 = wb['Раздел 2.1.1']
            self._fill_section2_1_1(ws_section2_1_1, data)
            
            wb.save(output_path)
            logger.info(f"Декларация сохранена: {output_path}")
            
        except Exception as e:
            logger.error(f"Ошибка генерации декларации: {e}")
            raise
    
    def _fill_title(self, ws, inn: str, name: str, oktmo: str, tax_code: str):
        """Заполняет титульный лист"""
        # ИНН по цифрам (ячейки G3-R3)
        for i, digit in enumerate(inn[:12]):
            ws.cell(row=3, column=7+i).value = digit
        
        # Наименование
        ws['B10'] = name
        
        # Код налогового органа
        ws['G16'] = tax_code
        
        # ОКТМО
        ws['G12'] = oktmo
    
    def _fill_section1_1(self, ws, data: Dict, oktmo: str):
        """Заполняет Раздел 1.1"""
        # ОКТМО для всех кварталов
        ws['B5'] = oktmo
        ws['B14'] = oktmo
        ws['B23'] = oktmo
        ws['B32'] = oktmo
        
        # Суммы к уплате
        ws['B9'] = round(data.get('to_pay_q1', 0), 2)
        ws['B18'] = round(data.get('to_pay_q2', 0), 2)
        ws['B27'] = round(data.get('to_pay_q3', 0), 2)
        ws['B36'] = round(data.get('to_pay_year', 0), 2)
    
    def _fill_section2_1_1(self, ws, data: Dict):
        """Заполняет Раздел 2.1.1"""
        # Признак налогоплательщика (2 = ИП без работников)
        ws['B5'] = 2
        
        # Доходы нарастающим итогом
        ws['B9'] = round(data.get('income_q1', 0), 2)
        ws['B13'] = round(data.get('income_q1', 0) + data.get('income_q2', 0), 2)
        ws['B17'] = round(data.get('income_q1', 0) + data.get('income_q2', 0) + data.get('income_q3', 0), 2)
        ws['B21'] = round(sum([data.get(f'income_{q}', 0) for q in ['q1', 'q2', 'q3', 'q4']]), 2)
        
        # Ставка 6%
        for row in [25, 29, 33, 37]:
            ws.cell(row=row, column=2).value = 6.0
        
        # Исчисленный налог
        ws['B41'] = round(data.get('tax_q1', 0), 2)
        ws['B45'] = round(data.get('tax_q2', 0), 2)
        ws['B49'] = round(data.get('tax_q3', 0), 2)
        ws['B53'] = round(data.get('tax_year', 0), 2)
        
        # Страховые взносы
        insurance = data.get('insurance', 0)
        ws['B57'] = round(insurance * 0.25, 2)
        ws['B61'] = round(insurance * 0.5, 2)
        ws['B65'] = round(insurance * 0.75, 2)
        ws['B69'] = round(insurance, 2)